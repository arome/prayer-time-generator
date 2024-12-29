import os
import openpyxl
import pandas as pd
import datetime as dt
import re

input_dir = "./input/"
input_files = os.listdir(input_dir)
filename = [s for s in input_files if "xlsx" in s][0]
print(f"Using {filename} as input file")
input_file_path = input_dir + filename
wb = openpyxl.load_workbook(input_file_path, data_only=True)
try:
    iqama_date_row = int(
        input('Enter the row# of the iqama word "date" (default 47): ')
    )
except ValueError:
    iqama_date_row = 47
    print("Defaulting to 47...")

current_year = dt.datetime.now().year
next_year = current_year + 1
try:
    year = int(input(f"Enter the year (default {next_year}): "))
except ValueError:
    year = next_year
    print(f"Defaulting to {next_year}...")
sheet = wb.active

def get_months_row():
    for row in range(sheet.max_row):
        for column in range(sheet.max_column):
            cell = sheet.cell(row + 1, column + 1)
            if cell.value == "January":
                return row + 1


def get_months_columns(months_row):
    months_columns = []
    for column in range(sheet.max_column):
        cell = sheet.cell(months_row, column + 1)
        if cell.value in months:
            months_columns.append(column + 1)
    return months_columns


months = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]
months_short = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]
months_row = get_months_row()
months_columns = get_months_columns(months_row)

prayer_times = dict()
df = pd.read_csv("input/moonode-example.csv")
# Ensure your 'date' column is of datetime type
df["date"] = df["date"].replace("2020", str(year), regex=True)

def adjust_time(time, cond=False):
    if cond and int(time.strftime("%H")) >= 9:
        return time
    result = dt.datetime.combine(dt.date.today(), time) + dt.timedelta(hours=12)
    return result.time()


def adjust_iqama_time(time, athan_time, convert=True, cond=False):
    if type(time) is str:
        if "+" in time:
            minutes = re.findall(r"\d+", time)[0]
            iqama_time = dt.datetime.combine(
                dt.date.today(), athan_time
            ) + dt.timedelta(minutes=int(minutes))
            return iqama_time.time()
        time = dt.time.fromisoformat(time.rjust(5, "0") + ":00")
    if convert:
        return adjust_time(time, cond)
    else:
        return time


website_time = []
for i, month_column in enumerate(months_columns):
    row = months_row + 1
    cell = sheet.cell(row, month_column)
    while cell.value:
        date_string = (
            str(i + 1).rjust(2, "0")
            + "/"
            + str(cell.value).replace("*", "").strip().rjust(2, "0")
        )
        prayer_times[date_string] = {
            "adhanFajr": sheet.cell(row, month_column + 1).value,
            "shourouk": sheet.cell(row, month_column + 2).value,
            "adhanDhuhr": adjust_time(sheet.cell(row, month_column + 3).value, True),
            "adhanAsr": adjust_time(sheet.cell(row, month_column + 4).value),
            "adhanMaghrib": adjust_time(sheet.cell(row, month_column + 5).value),
            "adhanIsha": adjust_time(sheet.cell(row, month_column + 6).value),
        }
        row = row + 1
        cell = sheet.cell(row, month_column)
    row = iqama_date_row + 1
    cell = sheet.cell(row, month_column + 1)
    while cell.value:
        if "To" in cell.value:
            from_date, to_date = cell.value.split(" To ")
        else: 
            from_date = to_date = cell.value
        for d in range(int(from_date), int(to_date) + 1):
            date_string = str(i + 1).rjust(2, "0") + "/" + str(d).rjust(2, "0")
            prayers = prayer_times[date_string]
            prayers["iqamaFajr"] = adjust_iqama_time(
                sheet.cell(row, month_column + 2).value, prayers["adhanFajr"], False
            )
            prayers["iqamaDhuhr"] = adjust_iqama_time(
                sheet.cell(row, month_column + 3).value,
                prayers["adhanDhuhr"],
                True,
                True,
            )
            prayers["iqamaAsr"] = adjust_iqama_time(
                sheet.cell(row, month_column + 4).value, prayers["adhanAsr"]
            )
            prayers["iqamaMaghrib"] = adjust_iqama_time(
                sheet.cell(row, month_column + 5).value, prayers["adhanMaghrib"]
            )
            prayers["iqamaIsha"] = adjust_iqama_time(
                sheet.cell(row, month_column + 6).value, prayers["adhanIsha"]
            )
            df.loc[
                df["date"] == date_string + "/" + str(year),
                [
                    "adhanFajr",
                    "iqamaFajr",
                    "shourouk",
                    "adhanDhuhr",
                    "iqamaDhuhr",
                    "adhanAsr",
                    "iqamaAsr",
                    "adhanMaghrib",
                    "iqamaMaghrib",
                    "adhanIsha",
                    "iqamaIsha",
                ],
            ] = [
                format(prayers["adhanFajr"], "%I:%M %p"),
                format(prayers["iqamaFajr"], "%I:%M %p"),
                format(prayers["shourouk"], "%I:%M %p"),
                format(prayers["adhanDhuhr"], "%I:%M %p"),
                format(prayers["iqamaDhuhr"], "%I:%M %p"),
                format(prayers["adhanAsr"], "%I:%M %p"),
                format(prayers["iqamaAsr"], "%I:%M %p"),
                format(prayers["adhanMaghrib"], "%I:%M %p"),
                format(prayers["iqamaMaghrib"], "%I:%M %p"),
                format(prayers["adhanIsha"], "%I:%M %p"),
                format(prayers["iqamaIsha"], "%I:%M %p"),
            ]
            website_time.append(
                f"{months_short[i]} {d}--{format(prayers['adhanFajr'], '%I:%M %p')}--{format(prayers['adhanDhuhr'], '%I:%M %p')}--{format(prayers['adhanAsr'], '%I:%M %p')}--{format(prayers['adhanMaghrib'], '%I:%M %p')}--{format(prayers['adhanIsha'], '%I:%M %p')}--{format(prayers['iqamaFajr'], '%I:%M %p')}--{format(prayers['iqamaDhuhr'], '%I:%M %p')}--{format(prayers['iqamaAsr'], '%I:%M %p')}--{format(prayers['iqamaMaghrib'], '%I:%M %p')}--{format(prayers['iqamaIsha'], '%I:%M %p')}"
            )
        row = row + 1
        cell = sheet.cell(row, month_column + 1)
print(df)
df.to_csv("output/moonode.csv", index=False)
print("outputting data to output/moonode.csv")

with open("output/website.txt", "w") as f:
    f.write("\n".join(website_time))
print("outputting data to output/website.txt")
